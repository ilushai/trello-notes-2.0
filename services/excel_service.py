# services/excel_service.py
import openpyxl
from openpyxl import Workbook
import os
from threading import Lock

EXCEL_FILE_PATH = "notes.xlsx"
# Создаем "замок" для предотвращения одновременной записи в файл из разных потоков
file_lock = Lock()

def add_note(text: str):
    """
    Добавляет текстовую заметку в файл notes.xlsx в первую свободную строку колонки A.
    """
    with file_lock:
        # Проверяем, существует ли файл
        if os.path.exists(EXCEL_FILE_PATH):
            # Если да, открываем его
            workbook = openpyxl.load_workbook(EXCEL_FILE_PATH)
            sheet = workbook.active
        else:
            # Если нет, создаем новый
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Notes"
            # Можно добавить заголовок
            sheet["A1"] = "Мои заметки"

        # Находим первую пустую строку в колонке A
        next_row = sheet.max_row + 1
        
        # Записываем текст в ячейку A<next_row>
        sheet.cell(row=next_row, column=1, value=text)
        
        # Сохраняем изменения
        workbook.save(EXCEL_FILE_PATH)